import os
import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tqdm import tqdm
from time import perf_counter
from collections import defaultdict

# Define base directory relative to script location
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_PATH = os.path.join(BASE_DIR, "ip", "jobs_input.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "op")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load URLs and Locations
df_urls = pd.read_excel(INPUT_PATH, usecols=["Location", "URL"])
df_urls = df_urls.dropna(subset=["URL"])

# Setup Selenium driver
driver = webdriver.Chrome()

def load_all_jobs(url):
    driver.get(url)
    time.sleep(3)

    try:
        while True:
            view_more = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'View More Jobs')]"))
            )
            view_more.click()
            time.sleep(3)
    except:
        pass

    try:
        show_all = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Show All')]"))
        )
        show_all.click()
        time.sleep(5)
    except:
        pass

    return BeautifulSoup(driver.page_source, "html.parser")

def parse_jobs(soup, source_url):
    jobs = []
    job_list_items = soup.select("ul#search-results-jobs > li")

    for li in job_list_items:
        try:
            title = li.find("h2").get_text(strip=True)
        except:
            title = ""

        try:
            location = li.find("span", class_="job-location location-test").get_text(strip=True)
        except:
            location = ""

        try:
            add_locs = li.find("span", class_="additional-locations-values")
            additional_locations = add_locs.get_text(strip=True) if add_locs else ""
        except:
            additional_locations = ""

        try:
            department = li.find("span", class_="job-categories").get_text(strip=True)
        except:
            department = ""

        jobs.append({
            "Job Role": title,
            "Location": location,
            "Additional Locations": additional_locations,
            "Vertical": department,
            "Source URL": source_url
        })

    return jobs

# Start execution timer
start_time = perf_counter()

# Scrape data
output_rows = []
summary = defaultdict(int)

for _, row in tqdm(df_urls.iterrows(), total=len(df_urls), desc="Scraping Locations"):
    location_label = row["Location"]
    url = row["URL"]
    soup = load_all_jobs(url)
    jobs = parse_jobs(soup, url)

    if jobs:
        # Add the location row as a heading
        output_rows.append({
            "Job Role": location_label,  # Remove asterisks, we'll apply bold formatting later
            "Location": "", "Additional Locations": "",
            "Vertical": "", "Source URL": ""
        })

        output_rows.extend(jobs)
        summary[location_label] += len(jobs)

        # Add 3 empty rows for spacing
        for _ in range(3):
            output_rows.append({
                "Job Role": "", "Location": "",
                "Additional Locations": "", "Vertical": "", "Source URL": ""
            })

driver.quit()

# Save raw Excel
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = os.path.join(OUTPUT_DIR, f"jobs_output_{timestamp}.xlsx")
df_output = pd.DataFrame(output_rows)
df_output.to_excel(output_file, index=False)

# Insert blank row and column spacing
wb = load_workbook(output_file)
ws = wb.active

# Insert blank row after header
ws.insert_rows(2)

# Insert a blank column between each original column except the last one
num_cols = ws.max_column
for col in reversed(range(2, num_cols + 1)):
    ws.insert_cols(col)

# Adjust column widths for better visibility
for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 25  # Set column width to 25 for broader format

# Apply bold formatting to location names
bold_font = Font(bold=True)
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value in df_urls["Location"].values:  # Check if the cell value matches a location
            cell.font = bold_font

wb.save(output_file)

# End execution timer
end_time = perf_counter()
duration = end_time - start_time

# Print summary
print(f"nScraping complete in {duration:.2f} seconds.")
print(f"Output saved to: {output_file}")
print("nSummary of Jobs Extracted:")
for loc, count in summary.items():
    print(f"  - {loc}: {count} jobs")
